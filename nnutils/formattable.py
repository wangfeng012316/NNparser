# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 18:10:10 2020

@author: ll

Modified on Tue Sep 8 10:13:22 2020
@Stephen Qian
"""

import  pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.formatting.rule import CellIsRule

def str2num(x):
    if x=='':
        return 0
    else:
        return int(x)

def SumAndFormat(paraout, df=None):
    workbook = load_workbook(filename=paraout)
    sheet = workbook['Details'] #.active
    sheet.delete_rows(idx=3)
    sheet.freeze_panes = "A3"
    workbook.save(paraout)
    if df is None:
        # parse the df from workbook
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)

    sizeO = df.loc[:,('Size of Parameters','Output')].apply(str2num)
    sizeW = df.loc[:,('Size of Parameters','Weight')].apply(str2num)
    opGEMM = df.loc[:,('Operation Summary','GEMM')].apply(str2num)

    # max size and ops:

    SumSheetGen(
        paraout,
        [
            ['Total Activations(MB):',sizeO.sum()/(1000**2)],
            ['Total Weights(MB):',sizeW.sum()/(1000**2)],
            ['Total GEMM (G_ops):',opGEMM.sum()/(1000**3)]
        ]
    )
    # set global font etc.
    GlobalFormat(paraout)
    FormatTable(
        paraout,
        [
            sizeO.max(),
            sizeW.max(),
            opGEMM.max()
        ]
    )

def SumSheetGen(paraout, summaryContent):
    workbook = load_workbook(filename=paraout)
    if "Summary" in set(workbook.sheetnames):
        sheet = workbook["Summary"]
    else:
        sheet = workbook.create_sheet("Summary")
    for i in range(len(summaryContent)):
        sheet["A{}".format(i+1)] = summaryContent[i][0]
        sheet["B{}".format(i+1)] = summaryContent[i][1]

    sheet.insert_rows(idx=1)
    summarystr = 'Model Statistics:'
    sheet['A1']= summarystr
    sheet.column_dimensions['A'].width = max(25,len(summarystr))
    workbook.save(paraout)

def FormatTable(paraout, maxVal, sheet_name='Details'):
    workbook = load_workbook(filename=paraout)
    sheet = workbook[sheet_name] #.active
    sheet.column_dimensions['A'].width = 1
    sheet.column_dimensions['B'].width = 12

    # row 0: Grey bkcolor, Bold font
    for cell in list(sum(list(sheet.rows)[:2], ())):
        cell.fill = PatternFill("solid", fgColor="00C0C0C0")
        cell.font = Font(name='Calibri',bold=True)
        if cell.value=='Output':
            so=cell.column_letter
        if cell.value=='Weight':
            sw=cell.column_letter
        if cell.value=='GEMM':
            og=cell.column_letter

    # Max activation row with red
    background = PatternFill(bgColor="00FF0000")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxVal[0])], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(so+'{}:'.format(sheet.min_row)+so+'{}'.format(sheet.max_row), myrule)

    # Max activation row with pink
    background = PatternFill(bgColor="00FF00FF")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxVal[1])], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(sw+'{}:'.format(sheet.min_row)+sw+'{}'.format(sheet.max_row), myrule)

    #  Max Ops Gemm row with green
    background = PatternFill(bgColor="0000FF00")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxVal[2])], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(og+'{}:'.format(sheet.min_row)+og+'{}'.format(sheet.max_row), myrule)

    # print(sheet.merged_cells)
    workbook.save(paraout)

def GlobalFormat(filename):
    workbook = load_workbook(filename=filename)
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        # traverse all content on this sheet
        header = True
        for row in sheet.rows:
            for cell in row:
                if header:
                    cell.font = Font(name='Calibri', bold=True)
                else:
                    cell.font = Font(name='Calibri', bold=False)
            header = False
    workbook.save(filename)